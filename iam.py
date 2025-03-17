from flask import Flask, request, jsonify, send_from_directory, session
import openpyxl
import os
import logging

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Important for sessions

logging.basicConfig(filename='iam.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class IAM:
    def __init__(self, excel_file="iam_data.xlsx"):
        self.excel_file = excel_file
        self.resources = {
            "report": "read_report",
            "dashboard": "view_dashboard",
            "settings": "manage_settings",
        }

    def _load_workbook(self):
        return openpyxl.load_workbook(self.excel_file)

    def _save_workbook(self, wb):
        wb.save(self.excel_file)

    def authenticate(self, username, password):
        wb = self._load_workbook()
        sheet = wb["Users"]

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == username and row[1].value == password:
                return True
        return False

    def authorize(self, username, resource):
        wb = self._load_workbook()
        users_sheet = wb["Users"]
        roles_sheet = wb["Roles"]

        user_role = None
        for row in users_sheet.iter_rows(min_row=2):
            if row[0].value == username:
                user_role = row[2].value
                break

        if not user_role:
            return False

        role_permissions = None
        for row in roles_sheet.iter_rows(min_row=2):
            if row[0].value == user_role:
                role_permissions = row[1].value
                break

        if not role_permissions:
            return False

        required_permission = self.resources.get(resource)
        return required_permission in role_permissions.split(",")

    def create_user(self, username, password, role):
        wb = self._load_workbook()
        sheet = wb["Users"]
        sheet.append([username, password, role])
        self._save_workbook(wb)
        logging.info(f"User created: username={username}, role={role}")

    def update_user(self, username, password, role):
        wb = self._load_workbook()
        sheet = wb["Users"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == username:
                row[1].value = password
                row[2].value = role
                self._save_workbook(wb)
                logging.info(f"User updated: username={username}, role={role}")
                return True
        return False

    def delete_user(self, username):
        wb = self._load_workbook()
        sheet = wb["Users"]
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == username:
                rows_to_delete.append(row[0].row)

        if rows_to_delete:
            for row_index in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_index)
            self._save_workbook(wb)
            logging.info(f"User deleted: username={username}")
            return True
        return False

    def create_role(self, role_name, permissions):
        wb = self._load_workbook()
        sheet = wb["Roles"]
        sheet.append([role_name, permissions])
        self._save_workbook(wb)
        logging.info(f"Role created: role_name={role_name}, permissions={permissions}")

    def update_role(self, role_name, permissions):
        wb = self._load_workbook()
        sheet = wb["Roles"]
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == role_name:
                row[1].value = permissions
                self._save_workbook(wb)
                logging.info(f"Role updated: role_name={role_name}, permissions={permissions}")
                return True
        return False

    def delete_role(self, role_name):
        wb = self._load_workbook()
        sheet = wb["Roles"]
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == role_name:
                rows_to_delete.append(row[0].row)

        if rows_to_delete:
            for row_index in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_index)
            self._save_workbook(wb)
            logging.info(f"Role deleted: role_name={role_name}")
            return True
        return False

iam = IAM()

def get_user_role(username):
    wb = iam._load_workbook()
    sheet = wb["Users"]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == username:
            return row[2].value
    return None

@app.route('/authenticate', methods=['POST'])
def authenticate():
    data = request.get_json()
    username = data['username']
    password = data['password']
    success = iam.authenticate(username, password)

    if success:
        logging.info(f"Authentication successful: username={username}")
        session['username'] = username
        session['authenticated'] = True
        user_role = get_user_role(username)
        session['role'] = user_role
    else:
        logging.warning(f"Authentication failed: username={username}")
        session.pop('username', None)
        session.pop('authenticated', None)
        session.pop('role', None)

    return jsonify({'success': success})

@app.route('/authorize', methods=['POST'])
def authorize():
    data = request.get_json()
    username = data['username']
    resource = data['resource']
    success = iam.authorize(username, resource)
    return jsonify({'success': success})

@app.route('/create_user', methods=['POST'])
def create_user():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        iam.create_user(data['username'], data['password'], data['role'])
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': 'Unauthorized'})

@app.route('/update_user', methods=['POST'])
def update_user():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        success = iam.update_user(data['username'], data['password'], data['role'])
        return jsonify({'success': success})
    else:
        return jsonify({'success': False, 'message': 'Unauthorized'})

@app.route('/delete_user', methods=['POST'])
def delete_user():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        success = iam.delete_user(data['username'])
        return jsonify({'success': success})
    else:
        return jsonify({'success': False, 'message': 'Unauthorized'})

@app.route('/create_role', methods=['POST'])
def create_role():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        iam.create_role(data['role_name'], data['permissions'])
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'message': 'Unauthorized'})

@app.route('/update_role', methods=['POST'])
def update_role():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        success = iam.update_role(data['role_name'], data['permissions'])
        return jsonify({'success': success})
    else:
        return jsonify({'success': False, 'message': 'Unauthorized'})

@app.route('/delete_role', methods=['POST'])
def delete_role():
    if session.get('authenticated') and session.get('role') == 'admin':
        data = request.get_json()
        success = iam.delete_role(data['role_name'])
        return jsonify
    
@app.route('/') #Add this route.
def index():
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), 'auth.html')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
